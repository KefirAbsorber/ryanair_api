from datetime import datetime
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import pathlib
import configparser

# get date for logs
now = datetime.now()

# files setup
path = pathlib.Path(__file__).parent.resolve()
# excel
excel_file = path.joinpath('flights.xlsx')
try:
    wb = load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()
ws = wb.active
# config
config = configparser.ConfigParser()
config.read("settings.ini")


# api request and formating
def find(frm, to, date_from, date_to):
    url: str = "https://services-api.ryanair.com/farfnd/3/oneWayFares?&departureAirportIataCode=" + frm + "&arrivalAirportIataCode=" + to + "&language=pl&limit=16&market=pl-pl&offset=0&outboundDepartureDateFrom=" + date_from + "&outboundDepartureDateTo=" + date_to
    response: requests.Response = requests.get(url)
    response: dict = response.json()
    if response["total"] == 0:
        print("No fligts found on query: " + frm + " "+ to + " " + date_from + " " + date_to)  # maybe add a log file later

        return -1
    else:
        flight_out_date = response["fares"][0]["outbound"]["departureDate"]
        flight_in_date = response["fares"][0]["outbound"]["arrivalDate"]
        cost = response["fares"][0]["summary"]["price"]["value"]

        result = [now, flight_out_date, flight_in_date, frm, to, cost]

        return result


# tracking
queries = config.sections()
if not queries:
    raise ValueError("No queries found, add using config")
for query in queries:
    flight = find(config[query]["frm"], config[query]["to"], config[query]["date_from"], config[query]["date_to"])
    if flight != -1:
        ws.append(flight)

# save to worksheet
wb.save(excel_file)
